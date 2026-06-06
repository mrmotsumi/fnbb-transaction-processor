# =============================================================================
# FNBB Bank Transaction Processor — main.py
#
# FastAPI application that parses FNBB CSV bank statements, classifies
# transactions as payments vs disbursements, publishes payments to a
# downstream PHP lending system, and generates Excel/PDF reports.
#
# All sensitive values loaded from environment variables.
# Never commit a .env file containing real credentials.
#
# Required .env keys:
#   SECRET_KEY              - App signing secret (32+ chars, random)
#   JWT_SECRET_KEY          - JWT signing secret (32+ chars, random)
#   DB_PASSWORD             - MySQL password
#   PHP_API_BASE_URL        - Downstream PHP system base URL
#   PHP_API_KEY             - PHP system API key
#   ADMIN_DEFAULT_PASSWORD  - Initial admin password (change on first login)
#   REPORT_ACCOUNT_NAME     - Account name shown in report headers
#   REPORT_ACCOUNT_NUMBER   - Account number shown in report headers
# =============================================================================

from contextlib import asynccontextmanager
import pathlib, shutil, tempfile
from fastapi import (
    FastAPI, File, Form, UploadFile, HTTPException,
    Depends, Query, status, Request, Response, BackgroundTasks, Body,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.httpsredirect import HTTPSRedirectMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer
from openpyxl import load_workbook
from pydantic import BaseModel, Field, validator
from pydantic_settings import BaseSettings
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any, Callable
from datetime import datetime, timedelta, timezone
import os, hashlib, json, re, io, math, logging, time, uuid, platform, asyncio

from dotenv import load_dotenv
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from sqlalchemy import case, func, text, inspect, Column, Integer, String, DateTime, Float, JSON, Text
from sqlalchemy.ext.declarative import declarative_base
from mangum import Mangum
from pydantic import StringConstraints, field_validator
from typing import Annotated
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from redis import asyncio as aioredis
from functools import wraps
import httpx, xlsxwriter, pandas as pd

from database import SessionLocal, engine, get_db
from model.models import Base, BankTransaction
from model.user_models import User, UserRole
from auth.auth_service import (
    AuthService, get_current_user, require_role,
    get_admin_user, get_collection_officer_user,
    get_accountant_user, get_any_authenticated_user,
    ACCESS_TOKEN_EXPIRE_MINUTES,
)
from schemas import (
    DatabasePlateExtractionResponse, DatabasePlateResult,
    UserCreate, UserUpdate, UserResponse, UserLogin,
    Token, PasswordChange, UserListResponse, UserRoleEnum,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)
load_dotenv()

# =============================================================================
# SETTINGS
# =============================================================================

class Settings(BaseSettings):
    environment: str = Field("production", alias="ENVIRONMENT")

    db_host:     str = Field("127.0.0.1", alias="DB_HOST")
    db_port:     str = Field("3306",      alias="DB_PORT")
    db_user:     str = Field("root",      alias="DB_USER")
    db_password: str = Field(...,         alias="DB_PASSWORD")
    db_name:     str = Field("transaction_db", alias="DB_NAME")

    # Secrets — required, no defaults
    secret_key:     str = Field(..., alias="SECRET_KEY")
    jwt_secret_key: str = Field(..., alias="JWT_SECRET_KEY")

    allowed_hosts: List[str] = Field(["*"], alias="ALLOWED_HOSTS")

    redis_url:               str  = Field("redis://localhost:6379", alias="REDIS_URL")
    cache_ttl_daily_summary: int  = Field(300,  alias="CACHE_TTL_DAILY_SUMMARY")
    cache_ttl_transactions:  int  = Field(60,   alias="CACHE_TTL_TRANSACTIONS")
    cache_ttl_reports:       int  = Field(3600, alias="CACHE_TTL_REPORTS")
    cache_ttl_users:         int  = Field(120,  alias="CACHE_TTL_USERS")
    cache_enabled:           bool = Field(True, alias="CACHE_ENABLED")

    refresh_token_expire_days: int = Field(7, alias="REFRESH_TOKEN_EXPIRE_DAYS")

    php_api_base_url:    str = Field(..., alias="PHP_API_BASE_URL")
    php_api_key:         str = Field(..., alias="PHP_API_KEY")
    php_api_timeout:     int = Field(30,  alias="PHP_API_TIMEOUT")
    php_api_max_retries: int = Field(3,   alias="PHP_API_MAX_RETRIES")
    php_api_retry_delay: int = Field(5,   alias="PHP_API_RETRY_DELAY")
    php_batch_size:      int = Field(50,  alias="PHP_BATCH_SIZE")

    # Initial admin — from env, never hardcoded
    admin_email:            str = Field("admin@example.com", alias="ADMIN_EMAIL")
    admin_default_password: str = Field(..., alias="ADMIN_DEFAULT_PASSWORD")

    # Report display info (PDF/Excel headers only — not for auth)
    report_account_name:   str = Field("Account Holder", alias="REPORT_ACCOUNT_NAME")
    report_account_number: str = Field("XXXXXXXXXXXX",   alias="REPORT_ACCOUNT_NUMBER")

    @property
    def database_url(self) -> str:
        return (
            f"mysql+pymysql://{self.db_user}:{self.db_password}"
            f"@{self.db_host}:{self.db_port}/{self.db_name}"
        )

    @field_validator("secret_key", "jwt_secret_key")
    @classmethod
    def validate_secrets(cls, v, info):
        if len(v) < 32:
            raise ValueError(
                f"{info.field_name} must be 32+ characters. "
                "Generate: python -c \"import secrets; print(secrets.token_hex(32))\""
            )
        return v

    @field_validator("environment")
    @classmethod
    def validate_env(cls, v):
        if v not in ["development", "staging", "production"]:
            raise ValueError("ENVIRONMENT must be development, staging, or production")
        return v

    @field_validator("redis_url")
    @classmethod
    def validate_redis(cls, v):
        if not v.startswith("redis://"):
            raise ValueError("REDIS_URL must start with redis://")
        return v

    class Config:
        env_file = ".env"
        extra    = "ignore"


settings      = Settings()
redis_client  = None
app_start_time = datetime.now(timezone.utc)

# =============================================================================
# PAYMENT PUBLISH LOG MODEL
# =============================================================================

class PaymentPublishLog(Base):
    __tablename__  = "payment_publish_logs"
    __table_args__ = {"extend_existing": True}

    id             = Column(Integer, primary_key=True, index=True)
    transaction_id = Column(Integer, index=True)
    transaction_date = Column(String(50))
    amount         = Column(Float)
    description    = Column(Text)
    account_name   = Column(String(255))
    account_number = Column(String(100))
    source_file    = Column(String(255))
    status         = Column(String(50), default="pending")
    attempt_count  = Column(Integer, default=0)
    last_attempt   = Column(DateTime, nullable=True)
    next_retry     = Column(DateTime, nullable=True)
    php_response       = Column(JSON, nullable=True)
    php_transaction_id = Column(String(255), nullable=True)
    error_message      = Column(Text, nullable=True)
    created_at   = Column(DateTime, default=datetime.utcnow)
    updated_at   = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    published_at = Column(DateTime, nullable=True)

    def to_dict(self):
        return {
            "id":                 self.id,
            "transaction_id":     self.transaction_id,
            "amount":             self.amount,
            "description":        (self.description or "")[:100],
            "status":             self.status,
            "attempt_count":      self.attempt_count,
            "php_transaction_id": self.php_transaction_id,
            "published_at":       self.published_at.isoformat() if self.published_at else None,
            "created_at":         self.created_at.isoformat()   if self.created_at   else None,
            "error_message":      self.error_message,
        }

# =============================================================================
# PYDANTIC SCHEMAS
# =============================================================================

class PaymentPublishResponse(BaseModel):
    success:            bool
    message:            str
    php_transaction_id: Optional[str]           = None
    error_code:         Optional[str]           = None
    details:            Optional[Dict[str, Any]] = None

class PublishStatusResponse(BaseModel):
    transaction_id:     int
    status:             str
    attempt_count:      int
    php_transaction_id: Optional[str] = None
    error_message:      Optional[str] = None
    last_attempt:       Optional[str] = None
    next_retry:         Optional[str] = None

class PHPWebhookPayload(BaseModel):
    php_transaction_id: str
    transaction_id:     int
    status:             str
    message:            Optional[str]           = None
    processed_at:       Optional[str]           = None
    details:            Optional[Dict[str, Any]] = None

class BatchPublishResponse(BaseModel):
    batch_id:           str
    total_transactions: int
    successful:         int
    failed:             int
    pending:            int
    results:            List[Dict[str, Any]]

# =============================================================================
# PAYMENT PUBLISHING SERVICE
# =============================================================================

class PaymentPublishingService:
    """Publishes classified payment transactions to the downstream PHP lending system.
    Uses exponential back-off retry and logs every attempt."""

    def __init__(self, db: Session, settings: Settings):
        self.db       = db
        self.settings = settings
        self.client   = httpx.AsyncClient(
            timeout=settings.php_api_timeout,
            headers={
                "Content-Type":    "application/json",
                "X-API-Key":       settings.php_api_key,
                "User-Agent":      "TransactionProcessor/2.0",
                "Accept":          "application/json",
                "Accept-Encoding": "gzip, deflate",
            },
            follow_redirects=True,
        )

    async def close(self):
        await self.client.aclose()

    async def publish_single_transaction(
        self,
        transaction: BankTransaction,
        max_retries: Optional[int] = None,
    ) -> PaymentPublishResponse:
        retries = max_retries or self.settings.php_api_max_retries
        attempt = 0

        log = self.db.query(PaymentPublishLog).filter(
            PaymentPublishLog.transaction_id == transaction.id
        ).first()

        if not log:
            log = PaymentPublishLog(
                transaction_id=transaction.id,
                transaction_date=transaction.date,
                amount=transaction.amount,
                description=transaction.description,
                account_name=transaction.account_name,
                account_number=transaction.account_number,
                source_file=transaction.source,
                status="pending",
            )
            self.db.add(log)
            self.db.commit()
            self.db.refresh(log)

        while attempt <= retries:
            attempt        += 1
            log.attempt_count = attempt
            log.last_attempt  = datetime.utcnow()
            log.status        = "attempting"
            self.db.commit()

            try:
                payload = {
                    "transaction_id":   transaction.id,
                    "transaction_date": transaction.date,
                    "amount":           float(transaction.amount),
                    "description":      transaction.description,
                    "account_name":     transaction.account_name,
                    "account_number":   transaction.account_number,
                    "source_file":      transaction.source,
                    "upload_date":      transaction.date_uploaded,
                    "is_payment":       transaction.is_payment,
                    "is_disbursement":  transaction.is_disbursement,
                    "timestamp":        datetime.utcnow().isoformat(),
                }

                logger.info(f"Attempt {attempt}: Publishing TX {transaction.id}")
                response      = await self.client.post(
                    f"{self.settings.php_api_base_url}/receive_payment.php", json=payload
                )
                response_text = response.text

                try:
                    data = response.json()
                    if response.status_code == 200 and data.get("success"):
                        log.status             = "success"
                        log.php_response       = data
                        log.php_transaction_id = data.get("php_transaction_id") or str(data.get("payment_id", ""))
                        log.published_at       = datetime.utcnow()
                        log.error_message      = None
                        self.db.commit()
                        logger.info(f"TX {transaction.id} published. PHP ID: {log.php_transaction_id}")
                        return PaymentPublishResponse(
                            success=True, message=data.get("message", "Published successfully"),
                            php_transaction_id=log.php_transaction_id, details=data,
                        )

                    if data.get("status") == "pending_review" or data.get("requires_review"):
                        log.status = "pending_review"; log.php_response = data
                        log.error_message = "Payment requires manual review"
                        self.db.commit()
                        return PaymentPublishResponse(
                            success=True, message="Payment requires manual review",
                            php_transaction_id=data.get("php_transaction_id"), details=data,
                        )

                    error_msg = data.get("message", f"HTTP {response.status_code}")
                    log.error_message = f"API Error: {error_msg}"

                except json.JSONDecodeError:
                    error_msg = ("PHP API returned HTML error page — check PHP error logs"
                                 if "<html" in response_text.lower()
                                 else "Invalid JSON response from PHP API")
                    log.error_message = error_msg

                # Schedule retry or mark failed
                if attempt <= retries:
                    delay = self.settings.php_api_retry_delay * attempt
                    log.next_retry = datetime.utcnow() + timedelta(seconds=delay)
                    log.status     = "retrying"
                    logger.warning(f"TX {transaction.id} attempt {attempt} failed, retry in {delay}s")
                else:
                    log.status = "failed"
                    logger.error(f"TX {transaction.id} failed after {retries} attempts")
                self.db.commit()

                if attempt <= retries:
                    await asyncio.sleep(self.settings.php_api_retry_delay * attempt)

            except httpx.TimeoutException:
                log.error_message = f"Timeout after {self.settings.php_api_timeout}s"
                log.status = "retrying" if attempt <= retries else "failed"
                if attempt <= retries:
                    delay = self.settings.php_api_retry_delay * attempt
                    log.next_retry = datetime.utcnow() + timedelta(seconds=delay)
                    await asyncio.sleep(delay)
                self.db.commit()

            except httpx.ConnectError as e:
                log.error_message = f"Connection error: {e}"
                log.status = "retrying" if attempt <= retries else "failed"
                if attempt <= retries:
                    delay = self.settings.php_api_retry_delay * attempt
                    log.next_retry = datetime.utcnow() + timedelta(seconds=delay)
                    await asyncio.sleep(delay)
                self.db.commit()

            except httpx.HTTPStatusError as e:
                log.error_message = f"HTTP {e.response.status_code}: {e}"
                log.status        = "failed"
                self.db.commit()
                return PaymentPublishResponse(
                    success=False, message=log.error_message,
                    error_code=f"HTTP_{e.response.status_code}",
                )

            except Exception as e:
                log.error_message = f"Unexpected error: {e}"
                log.status        = "failed"
                self.db.commit()
                logger.error(f"Unexpected error TX {transaction.id}: {e}", exc_info=True)
                return PaymentPublishResponse(success=False, message=log.error_message)

        return PaymentPublishResponse(success=False, message="Max retries exceeded")

# =============================================================================
# BACKGROUND TASKS
# =============================================================================

async def publish_payments_in_background(
    publish_id: str, transaction_ids: List[int], db: Session
):
    logger.info(f"Background publishing started: {publish_id}")
    service = PaymentPublishingService(db, settings)
    try:
        txs      = db.query(BankTransaction).filter(BankTransaction.id.in_(transaction_ids)).all()
        payments = [t for t in txs if t.is_payment]
        skipped  = len(txs) - len(payments)
        if skipped:
            logger.warning(f"Skipping {skipped} non-payment transactions")
        if not payments:
            logger.error(f"No valid payments for {publish_id}")
            return
        for i, p in enumerate(payments):
            r = await service.publish_single_transaction(p)
            logger.info(f"Published {i+1}/{len(payments)}: TX {p.id} — {r.message}")
            if i < len(payments) - 1:
                await asyncio.sleep(0.5)
        logger.info(f"Background publishing complete: {publish_id}")
    except Exception as e:
        logger.error(f"Background publishing failed {publish_id}: {e}", exc_info=True)
    finally:
        await service.close()

# =============================================================================
# STARTUP MIGRATIONS
# =============================================================================

def run_startup_migrations():
    try:
        inspector = inspect(engine)
        if "payment_publish_logs" not in inspector.get_table_names():
            logger.info("Migration: Creating payment_publish_logs")
            PaymentPublishLog.__table__.create(engine)
            logger.info("Migration complete")
    except Exception as e:
        logger.error(f"Migration error: {e}")

# =============================================================================
# REDIS
# =============================================================================

async def init_redis():
    global redis_client
    try:
        redis_client = await aioredis.from_url(
            settings.redis_url, encoding="utf-8", decode_responses=True
        )
        await redis_client.ping()
        logger.info("Redis connected")
    except Exception as e:
        logger.error(f"Redis failed: {e}")
        redis_client = None

@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("Starting up...")
    await init_redis()
    run_startup_migrations()
    with SessionLocal() as db:
        initial = create_initial_admin_user(db)
        if initial:
            print("\n" + "=" * 60)
            print("INITIAL ADMIN CREDENTIALS")
            print("=" * 60)
            print(f"Username: {initial.username}")
            print(f"Email:    {initial.email}")
            print(f"Password: [set via ADMIN_DEFAULT_PASSWORD env var]")
            print("=" * 60)
            print("IMPORTANT: Change password immediately after first login!")
            print("=" * 60 + "\n")
    logger.info("Startup complete")
    yield
    logger.info("Shutting down...")
    if redis_client:
        await redis_client.close()
        logger.info("Redis closed")

# =============================================================================
# APP
# =============================================================================

app = FastAPI(
    title="FNBB Transaction Processor API",
    description=(
        "Parses FNBB CSV bank statements, classifies transactions as payments "
        "vs disbursements, publishes payments to a downstream PHP lending system, "
        "and generates Excel/PDF reports."
    ),
    version="2.1.0",
    lifespan=lifespan,
    docs_url  ="/docs"  if settings.environment != "production" else None,
    redoc_url ="/redoc" if settings.environment != "production" else None,
    openapi_tags=[
        {"name": "Authentication",   "description": "Login and token management"},
        {"name": "User Management",  "description": "Admin user management (Admin only)"},
        {"name": "Transactions",     "description": "CSV upload and transaction queries"},
        {"name": "Payments",         "description": "Payment publishing to PHP system"},
        {"name": "Reports",          "description": "Reports and analytics"},
        {"name": "Plate Extraction", "description": "Vehicle registration plate matching"},
        {"name": "System",           "description": "Health checks and system status"},
        {"name": "Webhooks",         "description": "Webhook endpoints for external systems"},
    ],
)

# =============================================================================
# MIDDLEWARE
# =============================================================================

# CORS — restrict allow_origins in production via ALLOWED_HOSTS env var
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],    # Tighten in production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.add_middleware(GZipMiddleware, minimum_size=1000)

if settings.environment == "production":
    app.add_middleware(HTTPSRedirectMiddleware)
    app.add_middleware(TrustedHostMiddleware, allowed_hosts=settings.allowed_hosts)

limiter = Limiter(key_func=get_remote_address, default_limits=["100/hour", "20/minute"])
app.state.limiter = limiter
app.add_middleware(SlowAPIMiddleware)
security = HTTPBearer()

@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"]  = "nosniff"
    response.headers["X-Frame-Options"]          = "DENY"
    response.headers["X-XSS-Protection"]         = "1; mode=block"
    response.headers["Referrer-Policy"]           = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]        = "camera=(), microphone=(), geolocation=()"
    return response

@app.middleware("http")
async def validate_request_size(request: Request, call_next):
    max_size = 250 * 1024 * 1024
    if request.method in ["POST", "PUT"]:
        cl = request.headers.get("content-length")
        if cl and int(cl) > max_size:
            raise HTTPException(413, f"Request too large. Max: {max_size//(1024*1024)} MB")
    return await call_next(request)

@app.middleware("http")
async def log_requests(request: Request, call_next):
    start  = time.time()
    rid    = str(uuid.uuid4())
    logger.info(f"→ {rid}: {request.method} {request.url.path}")
    response = await call_next(request)
    ms     = (time.time() - start) * 1000
    response.headers["X-Request-ID"] = rid
    logger.info(f"← {rid}: {response.status_code} {ms:.1f}ms")
    return response

# =============================================================================
# CACHE UTILITIES
# =============================================================================

def cache_key_builder(func: Callable, *args, **kwargs) -> str:
    parts    = [func.__module__, func.__name__]
    filtered = {k: v for k, v in kwargs.items()
                if k not in ["db", "current_user", "request", "skip_cache"]}
    for a in args:
        parts.append(str(a.__dict__) if hasattr(a, "__dict__") else str(a))
    for k, v in sorted(filtered.items()):
        parts.append(f"{k}={v}")
    return f"cache:{func.__name__}:{hashlib.md5(':'.join(parts).encode()).hexdigest()}"

async def get_cached_response(key: str) -> Any:
    if not settings.cache_enabled or not redis_client:
        return None
    try:
        v = await redis_client.get(key)
        return json.loads(v) if v else None
    except Exception as e:
        logger.error(f"Cache get: {e}")
        return None

async def set_cached_response(key: str, data: Any, ttl: int) -> None:
    if not settings.cache_enabled or not redis_client:
        return
    try:
        await redis_client.setex(key, ttl, json.dumps(data, default=str))
    except Exception as e:
        logger.error(f"Cache set: {e}")

async def invalidate_cache_pattern(pattern: str) -> None:
    if not settings.cache_enabled or not redis_client:
        return
    try:
        keys = await redis_client.keys(pattern)
        if keys:
            await redis_client.delete(*keys)
    except Exception as e:
        logger.error(f"Cache invalidate: {e}")

async def invalidate_transaction_caches():
    for p in ["cache:*get_daily_summary*","cache:*get_transactions*",
              "cache:*get_*_report*","cache:*transactions*"]:
        await invalidate_cache_pattern(p)

async def invalidate_user_caches():
    for p in ["cache:*users*","cache:*get_users*"]:
        await invalidate_cache_pattern(p)

def cached_response(ttl: int = 60):
    def decorator(func):
        @wraps(func)
        async def wrapper(*args, **kwargs):
            if kwargs.get("skip_cache"):
                return await func(*args, **kwargs)
            try:
                ck = cache_key_builder(func, *args, **kwargs)
                c  = await get_cached_response(ck)
                if c is not None:
                    return c
                r = await func(*args, **kwargs)
                await set_cached_response(ck, r, ttl)
                return r
            except Exception as e:
                logger.warning(f"Cache fallback: {e}")
                return await func(*args, **kwargs)
        return wrapper
    return decorator

# =============================================================================
# HELPERS
# =============================================================================

class CSVValidationRules(BaseModel):
    max_file_size:      int       = Field(default=150 * 1024 * 1024)
    allowed_mime_types: List[str] = Field(default=["text/csv","application/vnd.ms-excel"])
    max_rows:           int       = Field(default=10000)

    @validator("max_file_size")
    def check_size(cls, v):
        if v > 300 * 1024 * 1024:
            raise ValueError("Max file size cannot exceed 300 MB")
        return v


def get_transactions_data(
    transaction_type: str,
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    upload_date: Optional[str] = None, transaction_date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    if transaction_type == "payments":
        q = db.query(BankTransaction).filter(BankTransaction.is_payment == True)
    elif transaction_type == "disbursements":
        q = db.query(BankTransaction).filter(BankTransaction.is_disbursement == True)
    else:
        q = db.query(BankTransaction)
    if start_date:        q = q.filter(BankTransaction.date_uploaded >= start_date)
    if end_date:          q = q.filter(BankTransaction.date_uploaded <= end_date)
    if upload_date:       q = q.filter(BankTransaction.date_uploaded == upload_date)
    if transaction_date:  q = q.filter(BankTransaction.date == transaction_date)
    return q.order_by(BankTransaction.date.desc()).all()


def generate_excel_report(transactions, report_type, filters):
    if not transactions:
        raise ValueError(f"No {report_type} transactions found")
    data  = []
    total = 0
    for t in transactions:
        data.append({
            "Transaction Date": t.date,
            "Upload Date":      t.date_uploaded,
            "Amount (BWP)":     f"{t.amount:,.2f}",
            "Description":      t.description,
            "Account Name":     t.account_name,
            "Source File":      t.source,
            "Uploaded At":      t.uploaded_at.strftime("%Y-%m-%d %H:%M:%S") if t.uploaded_at else "N/A",
        })
        total += t.amount
    df = pd.DataFrame(data)
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
        sn = f"{report_type.capitalize()} Report"
        df.to_excel(writer, sheet_name=sn, index=False, startrow=4)
        wb = writer.book; ws = writer.sheets[sn]; bold = wb.add_format({"bold": True})
        ws.write("A1", f"{report_type.upper()} REPORT", bold)
        ws.write("A2", f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        fi = []
        if filters.get("start_date"):       fi.append(f"From: {filters['start_date']}")
        if filters.get("end_date"):         fi.append(f"To: {filters['end_date']}")
        if filters.get("upload_date"):      fi.append(f"Upload: {filters['upload_date']}")
        if filters.get("transaction_date"): fi.append(f"Tx: {filters['transaction_date']}")
        if fi: ws.write("A3", f"Filters: {' | '.join(fi)}")
        ws.write(f"A{len(df)+7}", "SUMMARY", bold)
        ws.write(f"A{len(df)+8}", f"Total: {len(df)} transactions")
        ws.write(f"A{len(df)+9}", f"Total Amount: BWP {total:,.2f}")
        for col, w in [("A:A",15),("B:B",12),("C:C",15),("D:D",40),("E:E",20),("F:F",15),("G:G",20)]:
            ws.set_column(col, w)
    out.seek(0)
    return out


def generate_pdf_report(transactions, report_type, filters):
    if not transactions:
        raise ValueError(f"No {report_type} transactions found")
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4)
    els  = []
    st   = getSampleStyleSheet()
    ts   = ParagraphStyle("T", parent=st["Heading1"], fontSize=16, spaceAfter=30, alignment=1)
    els.append(Paragraph(f"{report_type.upper()} REPORT", ts))
    els.append(Spacer(1, 20))
    info = [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"]
    for k, l in [("start_date","From"),("end_date","To"),("upload_date","Upload"),("transaction_date","Tx Date")]:
        if filters.get(k): info.append(f"{l}: {filters[k]}")
    for i in info:
        els.append(Paragraph(i, st["Normal"]))
    els.append(Spacer(1, 20))
    total = sum(t.amount for t in transactions)
    st2   = Table([["Total Transactions:", str(len(transactions))],["Total Amount:", f"BWP {total:,.2f}"]],
                  colWidths=[2*inch, 2*inch])
    st2.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),
                              ("FONTSIZE",(0,0),(-1,-1),10),("BOTTOMPADDING",(0,0),(-1,-1),12)]))
    els += [st2, Spacer(1, 30), Paragraph("Transaction Details", st["Heading2"]), Spacer(1, 10)]
    td   = [["Date","Amount (BWP)","Description"]]
    for t in transactions:
        d = t.description[:47]+"..." if len(t.description) > 50 else t.description
        td.append([t.date, f"{t.amount:,.2f}", d])
    tbl = Table(td, colWidths=[1.5*inch, 1.5*inch, 4*inch])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),("TEXTCOLOR",(0,0),(-1,0),colors.whitesmoke),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),10),
        ("BOTTOMPADDING",(0,0),(-1,0),12),("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("FONTSIZE",(0,1),(-1,-1),8),("GRID",(0,0),(-1,-1),1,colors.black),
        ("ALIGN",(1,1),(1,-1),"RIGHT"),("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    els.append(tbl)
    doc.build(els)
    buf.seek(0)
    return buf


def _is_strong_password(p: str) -> bool:
    return (len(p) >= 8 and bool(re.search(r"[A-Z]",p)) and bool(re.search(r"[a-z]",p))
            and bool(re.search(r"\d",p)) and bool(re.search(r"[!@#$%^&*(),.?\":{}|<>]",p)))


def create_initial_admin_user(db: Session):
    """Create initial admin from environment config if no users exist."""
    try:
        if db.query(User).count() > 0:
            return None
        logger.info("No users found — creating initial admin from env config...")
        admin = User(
            username="admin",
            email=settings.admin_email,
            full_name="System Administrator",
            role=UserRole.ADMIN,
            hashed_password=User.get_password_hash(settings.admin_default_password),
        )
        db.add(admin); db.commit(); db.refresh(admin)
        logger.info(f"Initial admin created: {admin.username}")
        logger.warning("IMPORTANT: Change the admin password immediately!")
        return admin
    except Exception as e:
        logger.error(f"Initial admin creation failed: {e}")
        db.rollback()
        return None

# =============================================================================
# PLATE EXTRACTOR
# =============================================================================

class PlateExtractor:
    """Extracts Botswana vehicle registration plates from transaction description strings.
    Pattern: B followed by 3 digits and 2-3 letters (e.g. B123ABC)."""

    def __init__(self):
        self.plate_pattern = r"B\d{3}[A-C][A-Z]{2}"

    def extract_plate_number(self, reference: str) -> Optional[str]:
        if not isinstance(reference, str):
            return None
        m = re.findall(self.plate_pattern, reference.upper())
        return m[0] if m else None

    def extract_plates_from_db_transactions(self, transactions: List[BankTransaction]) -> List[Dict]:
        results = []
        for t in transactions:
            p = self.extract_plate_number(t.description)
            if p:
                results.append({
                    "transaction_id":   t.id,
                    "transaction_date": t.date,
                    "amount":           t.amount,
                    "reference":        t.description,
                    "plate_number":     p,
                    "upload_date":      t.date_uploaded,
                    "source":           t.source,
                    "transaction":      t,
                })
        return results

    async def search_excel_for_plates(
        self, excel_file_path: str, sheet_name: str,
        plate_column: str, plate_numbers: List[str],
        use_chunks: bool = True, chunksize: int = 10000,
    ) -> Dict[str, Dict]:
        plate_numbers = [p.upper() for p in plate_numbers if p]
        try:
            return await (self._search_excel_with_chunks(excel_file_path, sheet_name, plate_column, plate_numbers, chunksize)
                          if use_chunks
                          else self._search_excel_entire(excel_file_path, sheet_name, plate_column, plate_numbers))
        except Exception as e:
            logger.error(f"Excel search error: {e}")
            return {}

    async def _search_excel_with_chunks(
        self, path: str, sheet: str, col: str, plates: List[str], _chunksize: int
    ) -> Dict[str, Dict]:
        import openpyxl
        results = {}
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close(); return results
        ws         = wb[sheet]
        header_row = None
        col_idx    = None
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header_row = row; break
        if header_row:
            for i, c in enumerate(header_row):
                if c and str(c).strip() == col:
                    col_idx = i; break
        if col_idx is None:
            wb.close(); return results
        plate_set = set(plates)
        for rc, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            if rc % 10000 == 0:
                logger.info(f"Processed {rc} rows...")
            if len(row) > col_idx and row[col_idx]:
                p = str(row[col_idx]).upper().strip()
                if p in plate_set and p not in results:
                    results[p] = {
                        "found": True, "row_index": rc + 1,
                        "data": {str(header_row[i]): row[i] for i in range(min(len(row), len(header_row)))},
                        "full_match": True,
                    }
                    plate_set.discard(p)
                    if len(results) == len(plates): break
        wb.close()
        return results

    async def _search_excel_entire(
        self, path: str, sheet: str, col: str, plates: List[str]
    ) -> Dict[str, Dict]:
        results = {}
        df = pd.read_excel(path, sheet_name=sheet)
        if col in df.columns:
            df["_p"] = df[col].astype(str).str.upper().str.strip()
            for plate in plates:
                m = df[df["_p"] == plate]
                if not m.empty:
                    results[plate] = {"found": True, "row_index": int(m.index[0]),
                                      "data": m.iloc[0].to_dict(), "full_match": True}
        return results

# =============================================================================
# SYSTEM ENDPOINTS
# =============================================================================

@app.get("/health/", tags=["System"])
async def health_check(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        tables  = inspect(db.get_bind()).get_table_names()
        r_ok    = redis_client and await redis_client.ping()
        return {
            "status":        "healthy" if "bank_transactions" in tables else "degraded",
            "database":      "connected",
            "redis":         "connected" if r_ok else "disconnected",
            "publish_table": "payment_publish_logs" in tables,
            "timestamp":     datetime.now(timezone.utc).isoformat(),
            "environment":   settings.environment,
        }
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

@app.get("/system/status", tags=["System"])
async def system_status():
    return {
        "status": "healthy",
        "system": {"platform": platform.platform(), "python": platform.python_version()},
        "application": {
            "started_at":    app_start_time.isoformat(),
            "uptime":        str(datetime.now(timezone.utc) - app_start_time),
            "version":       app.version,
            "environment":   settings.environment,
            "cache_enabled": settings.cache_enabled,
        },
    }

@app.get("/setup/status", tags=["System"])
async def get_setup_status(db: Session = Depends(get_db)):
    count = db.query(User).count()
    return {"needs_setup": count == 0, "user_count": count}

@app.post("/setup/initial-admin", tags=["System"])
async def create_initial_admin(
    username: str = Query("admin"),
    email:    str = Query("admin@example.com"),
    full_name:str = Query("System Administrator"),
    password: str = Query(...),
    db: Session = Depends(get_db),
):
    """Create initial admin. Only works when no users exist.
    Prefer setting ADMIN_DEFAULT_PASSWORD in .env for fully automated setup."""
    if db.query(User).count() > 0:
        raise HTTPException(400, "Users already exist")
    if not _is_strong_password(password):
        raise HTTPException(400, "Password must be 8+ chars with upper, lower, digit, and special char")
    if db.query(User).filter((User.username == username) | (User.email == email)).first():
        raise HTTPException(400, "Username or email already taken")
    admin = User(username=username, email=email, full_name=full_name,
                 role=UserRole.ADMIN, hashed_password=User.get_password_hash(password))
    db.add(admin); db.commit(); db.refresh(admin)
    return {"message": "Initial admin created", "user": admin.to_dict(),
            "note": "Change password on first login"}

# =============================================================================
# AUTHENTICATION
# =============================================================================

@app.post("/auth/login", response_model=Token, tags=["Authentication"])
@limiter.limit("5/minute")
async def login(request: Request, credentials: UserLogin, db: Session = Depends(get_db)):
    user = AuthService.authenticate_user(db, credentials.username, credentials.password)
    if not user:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED,
                            "Incorrect username or password",
                            headers={"WWW-Authenticate": "Bearer"})
    if not user.is_active:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Account disabled")
    token = AuthService.create_access_token(
        data={"sub": user.username, "role": user.role.value},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES),
    )
    logger.info(f"User {user.username} logged in")
    return {"access_token": token, "token_type": "bearer",
            "expires_in": ACCESS_TOKEN_EXPIRE_MINUTES * 60, "user": user.to_dict()}

@app.get("/auth/me", response_model=UserResponse, tags=["Authentication"])
async def me(current_user: User = Depends(get_current_user)):
    return current_user.to_dict()

@app.put("/auth/change-password", tags=["Authentication"])
async def change_password(
    data: PasswordChange,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.verify_password(data.current_password):
        raise HTTPException(400, "Current password is incorrect")
    if not _is_strong_password(data.new_password):
        raise HTTPException(400, "Password does not meet complexity requirements")
    current_user.hashed_password = User.get_password_hash(data.new_password)
    db.commit(); await invalidate_user_caches()
    return {"message": "Password changed successfully"}

# =============================================================================
# USER MANAGEMENT (Admin only)
# =============================================================================

@app.post("/admin/users", response_model=UserResponse, tags=["User Management"])
async def create_user(
    user_data: UserCreate,
    current_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
):
    if not _is_strong_password(user_data.password):
        raise HTTPException(400, "Password does not meet complexity requirements")
    if db.query(User).filter((User.username == user_data.username) | (User.email == user_data.email)).first():
        raise HTTPException(400, "Username or email already registered")
    u = User(username=user_data.username, email=user_data.email, full_name=user_data.full_name,
             role=UserRole(user_data.role.value), hashed_password=User.get_password_hash(user_data.password))
    db.add(u); db.commit(); db.refresh(u)
    await invalidate_user_caches()
    logger.info(f"User {u.username} created by {current_user.username}")
    return u.to_dict()

@app.get("/admin/users", response_model=UserListResponse, tags=["User Management"])
@cached_response(ttl=settings.cache_ttl_users)
async def get_users(
    page: int = Query(1, ge=1), per_page: int = Query(10, ge=1, le=100),
    role: Optional[UserRoleEnum] = None, search: Optional[str] = None,
    current_user: User = Depends(get_admin_user), db: Session = Depends(get_db),
):
    q = db.query(User)
    if role:   q = q.filter(User.role == UserRole(role.value))
    if search:
        t = f"%{search}%"
        q = q.filter(User.username.ilike(t) | User.full_name.ilike(t) | User.email.ilike(t))
    total = q.count()
    users = q.order_by(User.created_at.desc()).offset((page-1)*per_page).limit(per_page).all()
    return {"users": [u.to_dict() for u in users], "total": total,
            "page": page, "per_page": per_page, "pages": math.ceil(total/per_page)}

@app.get("/admin/users/{user_id}", response_model=UserResponse, tags=["User Management"])
@cached_response(ttl=settings.cache_ttl_users)
async def get_user(user_id: int, current_user: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    return u.to_dict()

@app.delete("/admin/users/{user_id}", tags=["User Management"])
async def delete_user(
    user_id: int,
    current_user: User = Depends(get_admin_user), db: Session = Depends(get_db),
):
    if user_id == current_user.id: raise HTTPException(400, "Cannot delete your own account")
    u = db.query(User).filter(User.id == user_id).first()
    if not u: raise HTTPException(404, "User not found")
    db.delete(u); db.commit(); await invalidate_user_caches()
    return {"message": "User deleted successfully"}

# =============================================================================
# PAYMENT PUBLISHING
# =============================================================================

@app.post("/payments/publish", tags=["Payments"])
async def publish_payments(
    background_tasks: BackgroundTasks,
    transaction_ids: Optional[List[int]] = Query(None),
    upload_date: Optional[str] = Query(None, description="YYYY-MM-DD or DD/MM/YYYY"),
    start_date:  Optional[str] = Query(None),
    end_date:    Optional[str] = Query(None),
    async_mode:  bool = Query(True),
    current_user: User = Depends(get_accountant_user),
    db: Session = Depends(get_db),
):
    """Publish ONLY payment transactions (is_payment=True) to downstream PHP system.
    Disbursements are never published through this endpoint."""

    def parse_date(d):
        if not d: return None
        if re.match(r"^\d{4}-\d{2}-\d{2}$", d): return d
        if re.match(r"^\d{2}/\d{2}/\d{4}$", d):
            day, m, y = d.split("/"); return f"{y}-{m}-{day}"
        if re.match(r"^\d{2}-\d{2}-\d{4}$", d):
            day, m, y = d.split("-"); return f"{y}-{m}-{day}"
        return d

    pu = parse_date(upload_date); ps = parse_date(start_date); pe = parse_date(end_date)
    q  = db.query(BankTransaction).filter(BankTransaction.is_payment == True)
    if transaction_ids: q = q.filter(BankTransaction.id.in_(transaction_ids))
    if pu: q = q.filter(BankTransaction.date_uploaded == pu)
    if ps: q = q.filter(BankTransaction.date_uploaded >= ps)
    if pe: q = q.filter(BankTransaction.date_uploaded <= pe)
    payments = [p for p in q.all() if p.is_payment]

    if not payments:
        sample = [d[0] for d in db.query(BankTransaction.date_uploaded)
                  .filter(BankTransaction.is_payment == True).distinct().limit(5).all()]
        raise HTTPException(404, {"message": "No payments found",
                                  "sample_upload_dates": sample,
                                  "note": "Use YYYY-MM-DD format for dates"})

    already = db.query(PaymentPublishLog).filter(
        PaymentPublishLog.transaction_id.in_([p.id for p in payments]),
        PaymentPublishLog.status.in_(["success","confirmed"]),
    ).count()

    if async_mode:
        pid = str(uuid.uuid4())
        background_tasks.add_task(publish_payments_in_background, pid, [p.id for p in payments], db)
        return {"message": f"Publishing started: {len(payments)} payments",
                "publish_id": pid, "total": len(payments), "already_published": already}

    svc = PaymentPublishingService(db, settings)
    try:
        results = []; ok = 0; fail = 0
        for p in payments:
            r = await svc.publish_single_transaction(p)
            results.append({"transaction_id": p.id, "amount": p.amount,
                             "success": r.success, "message": r.message,
                             "php_transaction_id": r.php_transaction_id})
            if r.success: ok += 1
            else:         fail += 1
        return {"message": "Publishing complete", "total": len(payments),
                "successful": ok, "failed": fail, "results": results[:100]}
    finally:
        await svc.close()

@app.get("/payments/publish/status/{transaction_id}", tags=["Payments"])
async def get_publish_status(
    transaction_id: int,
    current_user: User = Depends(get_any_authenticated_user),
    db: Session = Depends(get_db),
):
    log = db.query(PaymentPublishLog).filter(
        PaymentPublishLog.transaction_id == transaction_id
    ).first()
    if not log:
        return {"transaction_id": transaction_id, "status": "not_published"}
    return PublishStatusResponse(
        transaction_id=log.transaction_id, status=log.status,
        attempt_count=log.attempt_count, php_transaction_id=log.php_transaction_id,
        error_message=log.error_message,
        last_attempt=log.last_attempt.isoformat() if log.last_attempt else None,
        next_retry=log.next_retry.isoformat()     if log.next_retry   else None,
    )

@app.get("/payments/publish/summary", tags=["Payments"])
@cached_response(ttl=60)
async def publish_summary(
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    current_user: User = Depends(get_accountant_user), db: Session = Depends(get_db),
):
    q = db.query(PaymentPublishLog)
    if start_date: q = q.filter(PaymentPublishLog.created_at >= start_date)
    if end_date:   q = q.filter(PaymentPublishLog.created_at <= end_date)
    total = q.count()
    ok    = q.filter(PaymentPublishLog.status == "success").count()
    fail  = q.filter(PaymentPublishLog.status == "failed").count()
    pend  = q.filter(PaymentPublishLog.status == "pending").count()
    conf  = q.filter(PaymentPublishLog.status == "confirmed").count()
    amt   = db.query(func.sum(PaymentPublishLog.amount)).filter(
                PaymentPublishLog.status == "success").scalar() or 0
    return {"summary": {"total": total, "successful": ok, "failed": fail,
                        "pending": pend, "confirmed": conf,
                        "success_rate": (ok/total*100) if total else 0,
                        "total_amount_published": float(amt)}}

@app.get("/payments/debug", tags=["Payments"])
async def debug_payments(current_user: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    total = db.query(BankTransaction).count()
    pmt   = db.query(BankTransaction).filter(BankTransaction.is_payment == True).count()
    disb  = db.query(BankTransaction).filter(BankTransaction.is_disbursement == True).count()
    dates = [d[0] for d in db.query(BankTransaction.date_uploaded)
             .filter(BankTransaction.is_payment == True)
             .distinct().order_by(BankTransaction.date_uploaded.desc()).limit(10).all()]
    return {"summary": {"total": total, "payments": pmt, "disbursements": disb},
            "sample_payment_upload_dates": dates}

@app.get("/payments/test-php-connection", tags=["Payments"])
async def test_php_connection(current_user: User = Depends(get_admin_user)):
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.get(f"{settings.php_api_base_url}/health",
                            headers={"X-API-Key": settings.php_api_key})
            return {"success": r.status_code == 200, "status_code": r.status_code}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/payments/by-plate/{plate_number}", tags=["Payments"])
async def payments_by_plate(
    plate_number: str,
    include_transactions: bool = Query(False),
    from_date: Optional[str] = None, to_date: Optional[str] = None,
    current_user: User = Depends(get_any_authenticated_user),
    db: Session = Depends(get_db),
):
    plate = plate_number.strip().upper().replace(" ","")
    q     = db.query(BankTransaction).filter(BankTransaction.is_payment == True)
    if from_date: q = q.filter(BankTransaction.date_uploaded >= from_date)
    if to_date:   q = q.filter(BankTransaction.date_uploaded <= to_date)
    ext = PlateExtractor(); matched = []; total = 0
    for p in q.all():
        ep = ext.extract_plate_number(p.description)
        if ep and ep.replace(" ","") == plate:
            if include_transactions:
                matched.append({"transaction_id": p.id, "date": p.date,
                                 "amount": float(p.amount), "description": p.description})
            total += p.amount
    return {"plate_number": plate_number.upper(), "has_payments": len(matched)>0,
            "total_paid": total, "payment_count": len(matched),
            "payments": matched if include_transactions else []}

@app.post("/payments/by-plates/batch", tags=["Payments"])
async def payments_by_plates_batch(
    plate_numbers: List[str] = Body(...),
    include_transactions: bool = Query(False),
    current_user: User = Depends(get_any_authenticated_user),
    db: Session = Depends(get_db),
):
    pmts = db.query(BankTransaction).filter(BankTransaction.is_payment == True).all()
    ext  = PlateExtractor(); results = {}
    for plate in plate_numbers:
        pc = plate.strip().upper().replace(" ",""); matched = []; total = 0
        for p in pmts:
            ep = ext.extract_plate_number(p.description)
            if ep and ep.replace(" ","") == pc:
                if include_transactions:
                    matched.append({"transaction_id": p.id, "date": p.date,
                                    "amount": float(p.amount), "description": p.description[:100]})
                total += p.amount
        results[plate.upper()] = {"has_payments": len(matched)>0, "total_paid": total,
                                   "payment_count": len(matched),
                                   "payments": matched if include_transactions else []}
    return {"success": True, "results": results, "total_plates_checked": len(plate_numbers)}

# =============================================================================
# WEBHOOKS
# =============================================================================

@app.post("/webhooks/php/confirmation", tags=["Webhooks"])
async def php_confirmation(
    payload: PHPWebhookPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    if request.headers.get("X-API-Key") != settings.php_api_key:
        raise HTTPException(401, "Invalid API key")
    log = (db.query(PaymentPublishLog)
           .filter(PaymentPublishLog.transaction_id == payload.transaction_id,
                   PaymentPublishLog.php_transaction_id == payload.php_transaction_id).first()
           or db.query(PaymentPublishLog)
              .filter(PaymentPublishLog.transaction_id == payload.transaction_id)
              .order_by(PaymentPublishLog.created_at.desc()).first())
    if log:
        if payload.status in ("received","processed"):
            log.status = "confirmed"
            log.php_response = {**(log.php_response or {}), "confirmation": payload.dict()}
        elif payload.status == "failed":
            log.status = "php_failed"; log.error_message = payload.message
        db.commit()
        return {"success": True, "message": "Confirmation received"}
    return {"success": False, "message": "Transaction not found"}

# =============================================================================
# TRANSACTION ENDPOINTS
# =============================================================================

@app.post("/upload-csv/", response_model=dict, tags=["Transactions"])
@limiter.limit("10/minute")
async def upload_csv(
    request: Request, background_tasks: BackgroundTasks,
    file: UploadFile = File(...), auto_publish: bool = Form(False),
    validation_rules: CSVValidationRules = Depends(lambda: CSVValidationRules()),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_accountant_user),
):
    """Upload FNBB CSV bank statement.
    Skips the first 4 header lines (FNBB format), classifies each transaction
    as payment (credit) or disbursement (debit), and deduplicates via SHA hash."""
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "File must be a CSV")
    try:
        contents = await file.read()
        if len(contents) > validation_rules.max_file_size:
            raise HTTPException(413, f"File too large. Max: {validation_rules.max_file_size//(1024*1024)} MB")
        content_str = None
        for enc in ["utf-8","latin-1","windows-1252"]:
            try: content_str = contents.decode(enc); break
            except UnicodeDecodeError: continue
        if not content_str:
            raise HTTPException(400, "Cannot decode file. Use UTF-8, Latin-1, or Windows-1252.")

        lines      = content_str.split("\n")
        data_lines = lines[4:]   # FNBB CSV — skip first 4 header lines
        transactions = []
        for line in data_lines:
            if not line.strip(): continue
            parts = line.split(",")
            if len(parts) >= 4:
                d = parts[0].strip(); a = parts[1].strip(); b = parts[2].strip()
                desc = ",".join(parts[3:]).strip().strip('"')
                if d and a and b:
                    transactions.append({"Date": d, "Amount": a, "Balance": b, "Description": desc})

        if len(transactions) > validation_rules.max_rows:
            raise HTTPException(400, f"Too many rows. Max: {validation_rules.max_rows}")

        df = pd.DataFrame(transactions)
        if df.empty:
            return {"message": "No valid transactions found", "processed_transactions": 0,
                    "duplicate_transactions": 0}

        df["Amount"]  = pd.to_numeric(df["Amount"].str.replace(",","").str.replace(" ",""), errors="coerce")
        df["Balance"] = pd.to_numeric(df["Balance"].str.replace(",","").str.replace(" ",""), errors="coerce")
        df            = df.dropna(subset=["Date","Amount","Balance"])

        processed = 0; duplicates = 0; errors = []
        upload_date = datetime.now(timezone.utc).date().isoformat()

        for idx, row in df.iterrows():
            try:
                amount = float(row["Amount"])
                tmp = BankTransaction(
                    date=str(row["Date"]), amount=abs(amount), balance=float(row["Balance"]),
                    description=str(row["Description"]),
                    account_name=settings.report_account_name,
                    account_number=settings.report_account_number,
                    source=file.filename, date_uploaded=upload_date,
                )
                tmp.set_transaction_flags(amount)
                tx_hash = tmp.generate_hash()
                if db.query(BankTransaction).filter(BankTransaction.transaction_hash == tx_hash).first():
                    duplicates += 1; continue
                tx = BankTransaction(
                    date=str(row["Date"]), amount=abs(amount), balance=float(row["Balance"]),
                    description=str(row["Description"]),
                    account_name=settings.report_account_name,
                    account_number=settings.report_account_number,
                    source=file.filename, transaction_hash=tx_hash, date_uploaded=upload_date,
                )
                tx.set_transaction_flags(amount)
                db.add(tx); processed += 1
            except Exception as e:
                errors.append(f"Row {idx}: {e}")

        db.commit()
        await invalidate_transaction_caches()

        if auto_publish and processed > 0:
            pmts = db.query(BankTransaction).filter(
                BankTransaction.source == file.filename, BankTransaction.is_payment == True
            ).all()
            if pmts:
                pid = str(uuid.uuid4())
                background_tasks.add_task(publish_payments_in_background, pid, [t.id for t in pmts], db)
                return {"message": "File processed. Payment publishing started.",
                        "processed_transactions": processed, "duplicate_transactions": duplicates,
                        "payments_to_publish": len(pmts), "publish_id": pid, "errors": errors}

        logger.info(f"CSV upload: {processed} processed, {duplicates} duplicates")
        return {"message": "File processed successfully",
                "processed_transactions": processed, "duplicate_transactions": duplicates,
                "upload_date": upload_date, "errors": errors, "total_errors": len(errors)}

    except HTTPException: raise
    except Exception as e:
        db.rollback(); logger.error(f"CSV error {file.filename}: {e}")
        raise HTTPException(500, f"Error processing file: {e}")

@app.get("/transactions/", tags=["Transactions"])
@cached_response(ttl=settings.cache_ttl_transactions)
async def get_transactions(
    upload_date: Optional[str] = None,
    skip: int = Query(0, ge=0), limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db),
):
    q = db.query(BankTransaction)
    if upload_date: q = q.filter(BankTransaction.date_uploaded == upload_date)
    return [{"id": t.id, "date": t.date, "amount": t.amount, "balance": t.balance,
             "description": t.description, "is_payment": t.is_payment,
             "is_disbursement": t.is_disbursement, "date_uploaded": t.date_uploaded,
             "source": t.source}
            for t in q.order_by(BankTransaction.date.desc()).offset(skip).limit(limit).all()]

@app.get("/upload-dates/", response_model=List[str], tags=["Transactions"])
@cached_response(ttl=settings.cache_ttl_transactions)
async def get_upload_dates(db: Session = Depends(get_db)):
    return [d[0] for d in db.query(BankTransaction.date_uploaded).distinct().all()]

# =============================================================================
# REPORTS
# =============================================================================

@app.get("/reports/daily-summary/", tags=["Reports"])
@cached_response(ttl=settings.cache_ttl_daily_summary)
async def daily_summary(
    start_date: Optional[str] = None, end_date: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(
        BankTransaction.date_uploaded,
        func.sum(case((BankTransaction.is_payment == True,      BankTransaction.amount), else_=0)).label("pmts"),
        func.sum(case((BankTransaction.is_disbursement == True, BankTransaction.amount), else_=0)).label("disbs"),
        func.count(case((BankTransaction.is_payment == True,      1))).label("pc"),
        func.count(case((BankTransaction.is_disbursement == True, 1))).label("dc"),
    ).group_by(BankTransaction.date_uploaded)
    if start_date: q = q.filter(BankTransaction.date_uploaded >= start_date)
    if end_date:   q = q.filter(BankTransaction.date_uploaded <= end_date)
    return [{"date": r.date_uploaded, "total_payments": float(r.pmts or 0),
             "total_disbursements": float(r.disbs or 0),
             "payment_count": r.pc or 0, "disbursement_count": r.dc or 0,
             "net_flow": float((r.pmts or 0)-(r.disbs or 0))}
            for r in q.order_by(BankTransaction.date_uploaded.desc()).all()]

@app.get("/reports/payments/", tags=["Reports"])
@cached_response(ttl=settings.cache_ttl_transactions)
async def payments_report(
    upload_date: Optional[str]=None, transaction_date: Optional[str]=None,
    skip: int=Query(0,ge=0), limit: int=Query(100,ge=1,le=1000),
    db: Session=Depends(get_db),
):
    q = db.query(BankTransaction).filter(BankTransaction.is_payment == True)
    if upload_date:      q = q.filter(BankTransaction.date_uploaded == upload_date)
    if transaction_date: q = q.filter(BankTransaction.date == transaction_date)
    return [{"id": p.id, "transaction_date": p.date, "upload_date": p.date_uploaded,
             "amount": p.amount, "description": p.description, "source": p.source}
            for p in q.order_by(BankTransaction.amount.desc()).offset(skip).limit(limit).all()]

@app.get("/reports/disbursements/", tags=["Reports"])
@cached_response(ttl=settings.cache_ttl_transactions)
async def disbursements_report(
    upload_date: Optional[str]=None, transaction_date: Optional[str]=None,
    skip: int=Query(0,ge=0), limit: int=Query(100,ge=1,le=1000),
    db: Session=Depends(get_db),
):
    q = db.query(BankTransaction).filter(BankTransaction.is_disbursement == True)
    if upload_date:      q = q.filter(BankTransaction.date_uploaded == upload_date)
    if transaction_date: q = q.filter(BankTransaction.date == transaction_date)
    return [{"id": d.id, "transaction_date": d.date, "upload_date": d.date_uploaded,
             "amount": d.amount, "description": d.description, "source": d.source}
            for d in q.order_by(BankTransaction.amount.desc()).offset(skip).limit(limit).all()]

# =============================================================================
# DOWNLOADS
# =============================================================================

def _sfx(s,e,u,t):
    if s and e: return f"_{s}_to_{e}"
    if u:       return f"_upload_{u}"
    if t:       return f"_tx_{t.replace('/','-')}"
    return ""

def _excel_response(output, filename):
    return StreamingResponse(io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})

def _pdf_response(buf, filename):
    return StreamingResponse(io.BytesIO(buf.read()), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})

@app.get("/download/payments/excel",       tags=["Reports"])
async def dl_payments_excel(s: Optional[str]=None,e: Optional[str]=None,u: Optional[str]=None,t: Optional[str]=None,db: Session=Depends(get_db)):
    f = {"start_date":s,"end_date":e,"upload_date":u,"transaction_date":t}
    return _excel_response(generate_excel_report(get_transactions_data("payments",s,e,u,t,db),"payments",f),
                           f"payments{_sfx(s,e,u,t)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.get("/download/disbursements/excel",  tags=["Reports"])
async def dl_disbursements_excel(s: Optional[str]=None,e: Optional[str]=None,u: Optional[str]=None,t: Optional[str]=None,db: Session=Depends(get_db)):
    f = {"start_date":s,"end_date":e,"upload_date":u,"transaction_date":t}
    return _excel_response(generate_excel_report(get_transactions_data("disbursements",s,e,u,t,db),"disbursements",f),
                           f"disbursements{_sfx(s,e,u,t)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

@app.get("/download/payments/pdf",         tags=["Reports"])
async def dl_payments_pdf(s: Optional[str]=None,e: Optional[str]=None,u: Optional[str]=None,t: Optional[str]=None,db: Session=Depends(get_db)):
    f = {"start_date":s,"end_date":e}
    return _pdf_response(generate_pdf_report(get_transactions_data("payments",s,e,u,t,db),"payments",f),
                         f"payments{_sfx(s,e,u,t)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

@app.get("/download/disbursements/pdf",    tags=["Reports"])
async def dl_disbursements_pdf(s: Optional[str]=None,e: Optional[str]=None,u: Optional[str]=None,t: Optional[str]=None,db: Session=Depends(get_db)):
    f = {"start_date":s,"end_date":e}
    return _pdf_response(generate_pdf_report(get_transactions_data("disbursements",s,e,u,t,db),"disbursements",f),
                         f"disbursements{_sfx(s,e,u,t)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

# =============================================================================
# PLATE EXTRACTION
# =============================================================================

@app.post("/extract-plates-from-db/", response_model=DatabasePlateExtractionResponse, tags=["Plate Extraction"])
@limiter.limit("5/minute")
async def extract_plates(
    request: Request,
    excel_file: UploadFile = File(...), sheet_name: str = Form(...),
    plate_column: str = Form("Registration"),
    start_date: Optional[str]=Form(None), end_date: Optional[str]=Form(None),
    transaction_type: Optional[str]=Form(None),
    chunked_reading: bool=Form(True), chunksize: int=Form(10000),
    current_user: User = Depends(get_any_authenticated_user),
    db: Session = Depends(get_db),
):
    temp_dir = None
    try:
        if not excel_file.filename.lower().endswith((".xlsx",".xls")):
            raise HTTPException(400, "Excel file must be .xlsx or .xls")
        q = db.query(BankTransaction)
        if start_date: q = q.filter(BankTransaction.date_uploaded >= start_date)
        if end_date:   q = q.filter(BankTransaction.date_uploaded <= end_date)
        if transaction_type == "payments":      q = q.filter(BankTransaction.is_payment == True)
        elif transaction_type == "disbursements": q = q.filter(BankTransaction.is_disbursement == True)
        txs = q.order_by(BankTransaction.date.desc()).all()
        if not txs: raise HTTPException(404, "No transactions found with given criteria")
        ext   = PlateExtractor()
        prs   = ext.extract_plates_from_db_transactions(txs)
        if not prs: raise HTTPException(404, "No plate numbers found in transaction references")
        plates = list(set(r["plate_number"] for r in prs))
        temp_dir   = tempfile.mkdtemp()
        excel_path = pathlib.Path(temp_dir) / "upload.xlsx"
        with open(excel_path,"wb") as f: f.write(await excel_file.read())
        wb = load_workbook(excel_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            avail = wb.sheetnames; wb.close()
            raise HTTPException(400, f"Sheet '{sheet_name}' not found. Available: {avail}")
        wb.close()
        sr = await ext.search_excel_for_plates(str(excel_path), sheet_name, plate_column, plates, chunked_reading)
        out = []; found = 0; total_amt = 0.0
        for pr in prs:
            p  = pr["plate_number"]; tx = pr["transaction"]
            item = DatabasePlateResult(
                transaction_id=tx.id, transaction_date=tx.date, amount=float(tx.amount),
                reference=tx.description[:100], plate_number=p, match_status="NOT_FOUND",
                excel_data=None, upload_date=tx.date_uploaded, source=tx.source,
            )
            if p in sr and sr[p].get("found"):
                found += 1; total_amt += tx.amount
                item.match_status = "FOUND"
                item.excel_data   = {k:v for k,v in sr[p]["data"].items() if k != "plate_clean"}
            out.append(item)
        return DatabasePlateExtractionResponse(
            message="Extraction complete",
            transactions_processed=len(txs), unique_plate_numbers=len(plates),
            plates_found=found, success_rate=(found/len(plates)*100) if plates else 0,
            results=out[:200], total_amount_matched=total_amt,
        )
    except HTTPException: raise
    except Exception as e:
        logger.error(f"Plate extraction error: {e}", exc_info=True)
        raise HTTPException(500, f"Internal error: {e}")
    finally:
        if temp_dir and pathlib.Path(temp_dir).exists():
            try: shutil.rmtree(temp_dir)
            except: pass

@app.get("/preview-plates-in-db/", tags=["Plate Extraction"])
@cached_response(ttl=60)
async def preview_plates(
    start_date: Optional[str]=None, end_date: Optional[str]=None,
    transaction_type: Optional[str]=None,
    limit: int=Query(100,ge=1,le=1000), skip: int=Query(0,ge=0),
    current_user: User=Depends(get_any_authenticated_user),
    db: Session=Depends(get_db),
):
    q = db.query(BankTransaction)
    if start_date: q = q.filter(BankTransaction.date_uploaded >= start_date)
    if end_date:   q = q.filter(BankTransaction.date_uploaded <= end_date)
    if transaction_type == "payments":      q = q.filter(BankTransaction.is_payment == True)
    elif transaction_type == "disbursements": q = q.filter(BankTransaction.is_disbursement == True)
    txs = q.order_by(BankTransaction.date.desc()).offset(skip).limit(limit).all()
    ext = PlateExtractor(); data = []
    for t in txs:
        p = ext.extract_plate_number(t.description)
        if p:
            data.append({"transaction_id": t.id, "transaction_date": t.date,
                          "upload_date": t.date_uploaded, "amount": float(t.amount),
                          "reference": t.description[:100], "plate_number": p,
                          "is_payment": t.is_payment})
    unique = list(set(d["plate_number"] for d in data))
    return {"total_transactions_queried": len(txs),
            "transactions_with_plates": len(data),
            "unique_plate_numbers": len(unique),
            "total_amount": sum(d["amount"] for d in data),
            "plate_numbers": unique, "transactions": data}

# =============================================================================
# ERROR HANDLERS
# =============================================================================

@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(status_code=429,
        content={"detail": "Too many requests", "retry_after": f"{exc.retry_after}s"},
        headers={"Retry-After": str(exc.retry_after)})

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    logger.warning(f"HTTP {exc.status_code}: {exc.detail} — {request.url.path}")
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled: {exc} — {request.url.path}", exc_info=True)
    return JSONResponse(status_code=500, content={"detail": "Internal server error"})

# =============================================================================
# AWS LAMBDA HANDLER
# =============================================================================

handler = Mangum(app)

# =============================================================================
# LOCAL ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000,
                log_level="info", reload=settings.environment == "development")
